# -*- coding: utf-8 -*-
"""
Análise de Impacto dos Resultados — DataSift
============================================

Avalia o impacto de repetir conjuntos de amostras (no mínimo 3 por teste). Cada
teste é um bloco com o seu Erro Total Máximo (ETM) e o seu Intervalo de
Referência (IR) — puxados automaticamente da base de dados pelo nome do teste —
e a sua própria tabela de amostras (código de barras, Resultado 1, Resultado 2).
É possível analisar vários testes ao mesmo tempo.

Para cada amostra decide se a diferença entre os dois resultados tem impacto
analítico (excede o ETM) e/ou clínico (muda a interpretação pelo IR).

Base de dados (colunas): Teste, Equipamento, ETM, IR.

Página do app DataSift (pasta ``pages/``). Também roda de forma independente com
``streamlit run pages/2_Analise_de_Impacto.py``.
"""

import io
import os
import re
from datetime import datetime
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import streamlit as st

# --------------------------------------------------------------------------- #
# Identidade visual (mesma paleta do DataSift)
# --------------------------------------------------------------------------- #
COLOR_PRIMARY = "#073B4C"
COLOR_SECONDARY = "#00E5FF"
COLOR_TERTIARY = "#118AB2"
COLOR_BG = "#F8F9FA"

try:
    st.set_page_config(page_title="DataSift · Impacto", page_icon="🎯", layout="wide")
except Exception:
    pass

st.markdown(
    f"""
    <style>
        .stApp {{ background-color: {COLOR_BG} !important; }}
        h1, h2, h3, h4 {{ color: {COLOR_PRIMARY} !important; font-weight: 800 !important; }}
        p, span, div[data-testid="stMarkdownContainer"], label {{ color: #212529 !important; }}
        div[data-testid="stMetricValue"] {{ color: {COLOR_PRIMARY} !important; }}
    </style>
    """,
    unsafe_allow_html=True,
)

# --------------------------------------------------------------------------- #
# Conformidade LGPD / GDPR — mesma tela do app.py; chave de sessão compartilhada
# --------------------------------------------------------------------------- #
GDPR_TERMS = """
This tool is designed to process and filter data from spreadsheets. The files you upload may contain sensitive personal data (such as full name, date of birth, national ID numbers, health information, etc.), the processing of which is regulated by data protection laws like the General Data Protection Regulation (GDPR or LGPD).

It is your sole responsibility to ensure that all data used in this tool complies with applicable data protection regulations. We strongly recommend that you only use previously anonymized data to protect the privacy of data subjects.

The responsibility for the nature of the processed data is exclusively yours.

To proceed, you must confirm that the data to be used has been properly handled and anonymized.
"""

if "lgpd_accepted" not in st.session_state:
    st.session_state.lgpd_accepted = False

if not st.session_state.lgpd_accepted:
    st.header("Terms of Use and Data Protection Compliance")
    st.markdown(GDPR_TERMS)
    accepted = st.checkbox("By checking this box, I confirm that the data provided is anonymized.")
    if st.button("Continue", type="primary", disabled=not accepted):
        st.session_state.lgpd_accepted = True
        st.rerun()
    st.stop()

APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


# --------------------------------------------------------------------------- #
# Funções auxiliares
# --------------------------------------------------------------------------- #
def normalizar_serie_numerica(serie: pd.Series) -> pd.Series:
    """Converte texto/misto em float, aceitando vírgula ou ponto como decimal."""
    def _conv(x):
        if pd.isna(x):
            return np.nan
        s = str(x).strip()
        if s == "" or s.lower() in ("nan", "none", "na", "n/a", "-", "--", "."):
            return np.nan
        s = s.replace("\xa0", "").replace(" ", "")
        s = re.sub(r"[^0-9,.\-+]", "", s)
        if s in ("", "+", "-", ".", ","):
            return np.nan
        has_dot, has_comma = "." in s, "," in s
        if has_dot and has_comma:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif has_comma:
            s = s.replace(",", ".")
        if s.count(".") > 1:
            partes = s.split(".")
            s = "".join(partes[:-1]) + "." + partes[-1]
        try:
            return float(s)
        except ValueError:
            return np.nan
    return serie.apply(_conv)


def parse_ref_range(txt):
    """Extrai (limite_inferior, limite_superior) de um IR ('136 - 145', '< 190', '> 40')."""
    if pd.isna(txt):
        return (None, None)
    s = str(txt).strip()
    if s == "":
        return (None, None)
    low = s.lower()
    nums = re.findall(r"\d+(?:[.,]\d+)?", s)
    vals = [float(n.replace(",", ".")) for n in nums]
    if not vals:
        return (None, None)
    if ("<" in s or "≤" in s or "menor" in low or "até" in low or "up to" in low):
        return (None, vals[-1])
    if (">" in s or "≥" in s or "maior" in low or "acima" in low):
        return (vals[0], None)
    if len(vals) >= 2:
        return (min(vals[0], vals[1]), max(vals[0], vals[1]))
    return (None, None)


def classificar_ref(valor, limite_inf, limite_sup):
    """Classifica em Baixo / Normal / Alto; '—' se não houver intervalo."""
    if pd.isna(valor):
        return "—"
    inf_ok = limite_inf is not None and pd.notna(limite_inf)
    sup_ok = limite_sup is not None and pd.notna(limite_sup)
    if not inf_ok and not sup_ok:
        return "—"
    if inf_ok and valor < limite_inf:
        return "Baixo"
    if sup_ok and valor > limite_sup:
        return "Alto"
    return "Normal"


def classificar_com_zc(valor, lo, hi, zc_lo, zc_hi):
    """
    Como classificar_ref, mas se o teste tiver zona cinza e o valor cair dentro
    dela (zc_lo <= valor <= zc_hi), a interpretação é 'Indeterminado'.
    """
    if pd.isna(valor):
        return "—"
    if zc_lo is not None and zc_hi is not None and zc_lo <= valor <= zc_hi:
        return "Indeterminado"
    return classificar_ref(valor, lo, hi)


_COLS_ESPERADAS = ("Teste", "ETM", "IR")


def _ler_tabela(conteudo: bytes, nome: str) -> pd.DataFrame:
    """
    Lê CSV ou Excel a partir dos bytes. Para CSV, testa combinações de
    separador/decimal/encoding e escolhe a primeira que traz as colunas
    esperadas — evita 'mojibake' (UTF-8 lido como latin-1) e separador errado.
    """
    nome = (nome or "").lower()
    if nome.endswith((".xlsx", ".xls")):
        return pd.read_excel(io.BytesIO(conteudo), engine="openpyxl")
    tentativas = [
        dict(sep=";", decimal=",", encoding="utf-8-sig"),
        dict(sep=";", decimal=",", encoding="latin-1"),
        dict(sep=",", decimal=".", encoding="utf-8-sig"),
        dict(sep=",", decimal=".", encoding="latin-1"),
    ]
    ultimo = None
    for t in tentativas:
        try:
            df = pd.read_csv(io.BytesIO(conteudo), engine="python", **t)
        except Exception:
            continue
        cols = [str(c).strip() for c in df.columns]
        df.columns = cols
        if all(e in cols for e in _COLS_ESPERADAS):
            return df
        ultimo = df
    return ultimo


@st.cache_data(show_spinner="Lendo base de dados...")
def carregar_base():
    """
    Carrega a base de dados da raiz do projeto (``Base de Dados.xlsx`` preferido,
    senão ``Base de Dados.csv``). Devolve (base_testes, equipamentos):
      - base_testes: DataFrame com Teste, ETM, IR (aba que contém essas colunas);
      - equipamentos: lista da aba **Equipamentos** (coluna **Equipamentos**);
        se não houver essa aba, cai para os valores únicos da coluna Equipamento.
    """
    xlsx = os.path.join(APP_DIR, "Base de Dados.xlsx")
    csv = os.path.join(APP_DIR, "Base de Dados.csv")
    base, equipamentos = None, []

    if os.path.exists(xlsx):
        try:
            abas = pd.read_excel(xlsx, sheet_name=None, engine="openpyxl")
        except Exception:
            abas = {}
        for _, d in abas.items():
            d.columns = [str(c).strip() for c in d.columns]
        # base de testes: 1ª aba que tenha Teste + ETM + IR (senão a 1ª aba)
        for _, d in abas.items():
            if all(c in d.columns for c in ("Teste", "ETM", "IR")):
                base = d
                break
        if base is None and abas:
            base = list(abas.values())[0]
        # equipamentos: aba "Equipamentos", coluna "Equipamentos"
        for nome_aba, d in abas.items():
            if str(nome_aba).strip().lower() == "equipamentos":
                col = "Equipamentos" if "Equipamentos" in d.columns else (
                    d.columns[0] if len(d.columns) else None)
                if col is not None:
                    equipamentos = sorted(d[col].dropna().astype(str).str.strip().unique().tolist())
                break
    elif os.path.exists(csv):
        with open(csv, "rb") as fh:
            base = _ler_tabela(fh.read(), "Base de Dados.csv")
        if base is not None:
            base.columns = [str(c).strip() for c in base.columns]

    # fallback dos equipamentos: coluna "Equipamento" da base de testes
    if not equipamentos and base is not None and "Equipamento" in base.columns:
        equipamentos = sorted(base["Equipamento"].dropna().astype(str).str.strip().unique().tolist())

    return base, equipamentos


def to_excel(df: pd.DataFrame, cols_2dec=None) -> bytes:
    """Exporta .xlsx centralizado, com autofit e 2 casas nas colunas indicadas."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    cols_2dec = set(cols_2dec or [])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Impacto")
        ws = writer.sheets["Impacto"]
        centro = Alignment(horizontal="center", vertical="center")
        n = len(df)
        for i, col in enumerate(df.columns, start=1):
            letra = get_column_letter(i)
            largura = max([len(str(col))] + [len(str(v)) for v in df[col].values])
            ws.column_dimensions[letra].width = min(largura + 2, 60)
            for r in range(1, n + 2):
                ws.cell(row=r, column=i).alignment = centro
            if col in cols_2dec:
                for r in range(2, n + 2):
                    ws.cell(row=r, column=i).number_format = "0.00"
    return output.getvalue()


def analisar_bloco(entrada: pd.DataFrame, teste, etm, ir_txt, lo, hi, zc_lo=None, zc_hi=None):
    """Valida e avalia as amostras de um teste. Devolve (df_resultado, n_validas)."""
    am = entrada.copy()
    am["Código de barras"] = am["Código de barras"].astype(str).str.strip()
    am["R1"] = normalizar_serie_numerica(am["Resultado 1"])
    am["R2"] = normalizar_serie_numerica(am["Resultado 2"])
    _bc = am["Código de barras"].str.lower()
    val = am[~_bc.isin(["", "nan", "none"])
             & am["R1"].notna() & am["R2"].notna() & (am["R1"] != 0)].reset_index(drop=True)
    if len(val) < 3:
        return None, len(val)

    val["Erro total %"] = np.abs((val["R1"] / val["R2"]) - 1) * 100
    val["Excede ETM"] = val["Erro total %"].abs() > etm if etm is not None else False
    val["Interpretação R1"] = val["R1"].apply(lambda v: classificar_com_zc(v, lo, hi, zc_lo, zc_hi))
    val["Interpretação R2"] = val["R2"].apply(lambda v: classificar_com_zc(v, lo, hi, zc_lo, zc_hi))
    val["Mudou interpretação"] = ((val["Interpretação R1"] != val["Interpretação R2"])
                                  & (val["Interpretação R1"] != "—") & (val["Interpretação R2"] != "—"))
    _exc = np.asarray(val["Excede ETM"], dtype=bool)
    _mud = np.asarray(val["Mudou interpretação"], dtype=bool)
    val["Impacto"] = np.select(
        [_exc & _mud, _exc & ~_mud, ~_exc & _mud],
        ["Erro total e Interpretação discordantes, realizar análise crítica",
         "Erro total discordante, realizar análise crítica",
         "Interpretação discordante, realizar análise crítica"],
        default="Sem impacto")
    val.insert(0, "Teste", teste)
    val.insert(1, "ETM (%)", etm)
    val.insert(2, "IR", ir_txt)
    return val, len(val)


def gerar_pdf(detalhe: pd.DataFrame, equipamento: str, operador: str, data_problema: str) -> bytes:
    """
    Gera um PDF (A4 paisagem) com o 'Detalhe por amostra', pronto para assinatura/
    auditoria: **texto completo** na coluna Impacto, **quebra automática de linha**
    e **autofit** de linhas e colunas. A coluna Impacto sai colorida (verde/vermelho,
    como no app). Cabeçalho com equipamento, operador e data/hora. Usa reportlab.
    """
    from datetime import datetime
    from zoneinfo import ZoneInfo
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

    df = detalhe.copy()
    if "Erro total %" in df.columns:
        df["Erro total %"] = pd.to_numeric(df["Erro total %"], errors="coerce").map(
            lambda v: "" if pd.isna(v) else f"{v:.2f}")
    for c in ("Excede ETM", "Mudou interpretação"):
        if c in df.columns:
            df[c] = df[c].map(lambda v: "Sim" if bool(v) else "Não")
    df = df.astype(str)
    colunas = list(df.columns)

    ss = getSampleStyleSheet()
    st_titulo = ParagraphStyle("titulo", parent=ss["Title"], fontSize=15, alignment=TA_LEFT,
                               textColor=colors.HexColor("#073B4C"), spaceAfter=2)
    st_sub = ParagraphStyle("sub", parent=ss["Normal"], fontSize=9,
                            textColor=colors.HexColor("#333333"), spaceAfter=10)
    st_head = ParagraphStyle("head", parent=ss["Normal"], fontName="Helvetica-Bold",
                             fontSize=8, leading=10, alignment=TA_CENTER, textColor=colors.white)
    st_cel = ParagraphStyle("cel", parent=ss["Normal"], fontSize=8, leading=10, alignment=TA_CENTER)
    st_verde = ParagraphStyle("verde", parent=st_cel, textColor=colors.HexColor("#0F5132"))
    st_vermelho = ParagraphStyle("vermelho", parent=st_cel, fontName="Helvetica-Bold",
                                 textColor=colors.HexColor("#9B1C1C"))

    # Células como Paragraph -> quebra automática de linha; Impacto colorido pelo valor.
    linhas = [[Paragraph(str(c), st_head) for c in colunas]]
    for _, row in df.iterrows():
        cel = []
        for c in colunas:
            v = str(row[c])
            if c == "Impacto":
                cel.append(Paragraph(v, st_verde if v == "Sem impacto" else st_vermelho))
            else:
                cel.append(Paragraph(v, st_cel))
        linhas.append(cel)

    # Autofit das colunas: largura proporcional ao maior conteúdo, com teto (força a quebra).
    pagina = landscape(A4)
    util = pagina[0] - 20 * mm
    natural = {c: max([len(str(c))] + [len(str(v)) for v in df[c].values]) for c in colunas}
    TETO = 26
    peso = [min(natural[c], TETO) for c in colunas]
    larguras = [util * p / sum(peso) for p in peso]

    tab = Table(linhas, colWidths=larguras, repeatRows=1)   # repete o cabeçalho a cada página
    estilo = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#073B4C")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
    ]
    if "Impacto" in colunas:
        ci = colunas.index("Impacto")
        for i, v in enumerate(df["Impacto"].tolist(), start=1):
            cor = colors.HexColor("#E7F6EC") if v == "Sem impacto" else colors.HexColor("#FFE3E3")
            estilo.append(("BACKGROUND", (ci, i), (ci, i), cor))
    tab.setStyle(TableStyle(estilo))

    cab = (f"Equipamento: {equipamento} &nbsp;&nbsp;|&nbsp;&nbsp; Operador: {operador or '—'} "
           f"&nbsp;&nbsp;|&nbsp;&nbsp; Data do problema: {data_problema or '—'}"
           f"&nbsp;&nbsp;|&nbsp;&nbsp; Relatório gerado em "
           f"{datetime.now(ZoneInfo('America/Sao_Paulo')):%d/%m/%Y %H:%M} ")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagina, leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=12 * mm, bottomMargin=10 * mm, title="Análise de Impacto")
    doc.build([Paragraph("Análise de Impacto — Detalhe por amostra", st_titulo),
               Paragraph(cab, st_sub), tab])
    return buf.getvalue()


# =========================================================================== #
#                                INTERFACE
# =========================================================================== #
st.markdown("## 🎯 Análise de Impacto dos Resultados")
st.caption(
    "Avalia o impacto de repetir amostras (mínimo 3 por teste) — em um ou vários testes ao "
    "mesmo tempo. O Erro Total Máximo (ETM) e o Intervalo de Referência (IR) são puxados "
    "automaticamente da base de dados pelo nome do teste."
)

# ---- 1 · Operador responsável --------------------------------------------- #
with st.container(border=True):
    st.markdown("### 1 · Operador responsável")
    operador = st.text_input("Nome do operador responsável", key="operador",
                             placeholder="Digite o nome do operador responsável pela análise")

# ---- 2 · Data do problema ------------------------------------------------- #
with st.container(border=True):
    st.markdown("### 2 · Data do problema")
    data_problema = st.date_input("Data do problema identificado", value=None,
                                  format="DD/MM/YYYY", key="data_problema")

base, equipamentos = carregar_base()

if base is None or base.empty:
    st.error("A base de dados (`Base de Dados.xlsx` ou `Base de Dados.csv`, com as colunas "
             "Teste, ETM, IR) não foi encontrada na raiz do projeto.")
    st.stop()

faltando = [c for c in ["Teste", "ETM", "IR"] if c not in base.columns]
if faltando:
    st.error(f"A base de dados não tem a(s) coluna(s): {', '.join(faltando)}. "
             f"Colunas encontradas: {', '.join(map(str, base.columns))}.")
    st.stop()

if not equipamentos:
    st.error("Nenhum equipamento encontrado. Informe-os na aba **Equipamentos** (coluna "
             "**Equipamentos**) do `Base de Dados.xlsx`, ou na coluna **Equipamento** da base.")
    st.stop()

base = base.copy()
base["_ETM_num"] = normalizar_serie_numerica(base["ETM"])
# No Excel o ETM costuma ser guardado como fração (0,15 = 15%, célula formatada como %),
# e o pandas lê o valor bruto (0,15). Se a coluna inteira estiver em fração (todos <= 1),
# convertemos para porcentagem (x100). Valores já em % (ex.: 15) ficam como estão.
_etm_ok = base["_ETM_num"].dropna()
if len(_etm_ok) and _etm_ok.max() <= 1:
    base["_ETM_num"] = base["_ETM_num"] * 100
testes = sorted(base["Teste"].dropna().astype(str).str.strip().unique().tolist())


def lookup_teste(teste):
    """
    Devolve (etm, ir_txt, lo, hi, zc_txt, zc_lo, zc_hi) da base (1ª linha do teste).
    zc_* vem da coluna 'Zona cinza' (quando existe e está preenchida para o teste).
    """
    linha = base[base["Teste"].astype(str).str.strip() == str(teste)]
    etm = float(linha["_ETM_num"].dropna().iloc[0]) if not linha["_ETM_num"].dropna().empty else None
    ir_txt = str(linha["IR"].dropna().iloc[0]) if not linha["IR"].dropna().empty else ""
    lo, hi = parse_ref_range(ir_txt)
    zc_txt = ""
    if "Zona cinza" in base.columns and not linha["Zona cinza"].dropna().empty:
        zc_txt = str(linha["Zona cinza"].dropna().iloc[0]).strip().lstrip("'")
    zc_lo, zc_hi = parse_ref_range(zc_txt) if zc_txt else (None, None)
    return etm, ir_txt, lo, hi, zc_txt, zc_lo, zc_hi


# Perfis pré-configurados: ao escolher, o app já carrega os testes do perfil.
PERFIS = {
    "Perfil Hemograma": ["WBC", "RBC", "HGB", "HCT", "VCM", "HCM", "CHCM", "PLT", "RDW", "VPM",
                         "NEA", "LYA", "MOA", "EOA", "BAA"],
}


def _match_teste(nome, testes):
    """Acha o nome do teste na base que corresponde (ignorando maiúsc./minúsc. e espaços)."""
    alvo = str(nome).strip().lower()
    for t in testes:
        if str(t).strip().lower() == alvo:
            return t
    return None


def _cod_vazio(v) -> bool:
    """True quando o código de barras está em branco/ausente."""
    return str(v).strip().lower() in ("", "nan", "none", "na", "n/a")


# Estado dos blocos de teste (um id por bloco)
if "imp_blocos" not in st.session_state:
    st.session_state.imp_blocos = [1]
    st.session_state.imp_next = 2

# ---- 3 · Equipamento e testes --------------------------------------------- #
with st.container(border=True):
    st.markdown("### 3 · Equipamento e testes")
    equip_sel = st.selectbox("Equipamento (vale para todos os testes abaixo)",
                             equipamentos, index=0 if equipamentos else None)

    perfil_sel = st.selectbox(
        "Perfil (opcional) — ao escolher, já carrega todos os testes do perfil",
        ["(nenhum)"] + list(PERFIS.keys()), key="perfil_sel")
    if perfil_sel == "(nenhum)":
        if st.session_state.get("_perfil_aplicado"):
            # Voltou para "(nenhum)": desfaz o perfil e restaura os testes que
            # existiam antes de aplicá-lo (limpando os blocos criados pelo perfil).
            for bid in list(st.session_state.imp_blocos):
                st.session_state.pop(f"teste_{bid}", None)
                for k in [k for k in list(st.session_state.keys())
                          if k == f"am_{bid}" or k.startswith(f"am_{bid}__")]:
                    st.session_state.pop(k, None)
            bkp = (st.session_state.get("_perfil_backup")
                   or {"imp_blocos": [1], "imp_next": 2, "estados": {}})
            st.session_state.imp_blocos = list(bkp["imp_blocos"])
            st.session_state.imp_next = bkp["imp_next"]
            # Restaura os testes escolhidos nos blocos que existiam antes do perfil
            # (o Streamlit descarta o estado de widgets que deixam de ser renderizados).
            for k, v in bkp.get("estados", {}).items():
                st.session_state[k] = v
            st.session_state["_perfil_res"] = {}
            st.session_state["_perfil_aplicado"] = None
            st.session_state["_perfil_faltantes"] = []
            st.rerun()
        st.session_state["_perfil_aplicado"] = None
        st.session_state["_perfil_faltantes"] = []
    elif st.session_state.get("_perfil_aplicado") != perfil_sel:
        if not st.session_state.get("_perfil_aplicado"):
            # Guarda o estado atual (antes do perfil) para restaurar ao voltar a "(nenhum)":
            # ids dos blocos, próximo id e o teste escolhido em cada bloco. (O conteúdo
            # das tabelas de amostras não pode ser reatribuído via session_state — regra
            # do Streamlit para o data_editor — então restauramos a estrutura e os testes.)
            estados = {f"teste_{bid}": st.session_state[f"teste_{bid}"]
                       for bid in st.session_state.imp_blocos
                       if f"teste_{bid}" in st.session_state}
            st.session_state["_perfil_backup"] = {
                "imp_blocos": list(st.session_state.imp_blocos),
                "imp_next": st.session_state.imp_next,
                "estados": estados,
            }
        achados, faltantes = [], []
        for nome in PERFIS[perfil_sel]:
            m = _match_teste(nome, testes)
            (achados if m else faltantes).append(m or nome)
        if achados:
            ids = list(range(st.session_state.imp_next,
                             st.session_state.imp_next + len(achados)))
            for bid, t in zip(ids, achados):
                st.session_state[f"teste_{bid}"] = t
            st.session_state.imp_blocos = ids
            st.session_state.imp_next += len(achados)
        st.session_state["_perfil_aplicado"] = perfil_sel
        st.session_state["_perfil_faltantes"] = faltantes
        st.rerun()
    if st.session_state.get("_perfil_faltantes"):
        st.warning("Testes do perfil não encontrados na base (confira os nomes): "
                   + ", ".join(st.session_state["_perfil_faltantes"]))

    st.caption("Cada bloco abaixo é um **teste**, com o seu ETM e IR puxados da base e a sua "
               "própria tabela de amostras (mínimo 3). Use o ➕ da tabela para mais linhas e o "
               "botão **Adicionar teste** para mais testes.")

# Com um perfil ativo, o 1º bloco (RBC) é o "dono" dos códigos de barras: o que for
# digitado nele é replicado (só leitura) para os demais testes do perfil. Assim o
# usuário digita os códigos uma única vez e só preenche os resultados dos outros testes.
perfil_ativo = bool(st.session_state.get("_perfil_aplicado"))
blocos_ids = list(st.session_state.imp_blocos)
codigos_master = None   # códigos do 1º bloco, preenchidos ao renderá-lo
nome_master = "1º teste"  # nome do 1º teste (dono dos códigos), para as legendas

blocos_dados = []   # (teste, etm, ir_txt, lo, hi, zc_lo, zc_hi, entrada_df)
for pos, bid in enumerate(blocos_ids):
    eh_master = (pos == 0)
    replicar = perfil_ativo and not eh_master
    with st.container(border=True):
        top = st.columns([5, 1])
        with top[0]:
            _tkey = f"teste_{bid}"
            if _tkey in st.session_state:
                teste_sel = st.selectbox("Nome do teste", testes, key=_tkey)
            else:
                teste_sel = st.selectbox("Nome do teste", testes,
                                         index=0 if testes else None, key=_tkey)
        with top[1]:
            st.markdown("<div style='height:1.75rem'></div>", unsafe_allow_html=True)
            if len(blocos_ids) > 1 and st.button("🗑️ Remover", key=f"del_{bid}"):
                st.session_state.imp_blocos.remove(bid)
                st.rerun()

        etm, ir_txt, lo, hi, zc_txt, zc_lo, zc_hi = lookup_teste(teste_sel)
        if zc_txt:
            cE = st.columns(3)
            cE[2].metric("Zona cinza (indeterminado)", zc_txt)
        else:
            cE = st.columns(2)
        cE[0].metric("Erro Total Máximo (ETM)", f"{etm:.2f} %" if etm is not None else "—")
        cE[1].metric("Intervalo de Referência (IR)", ir_txt if ir_txt else "—")

        if replicar:
            # Códigos vêm do 1º bloco (só leitura); resultados guardados por código.
            cods = ["" if _cod_vazio(c) else str(c).strip()
                    for c in (codigos_master if codigos_master is not None else ["", "", ""])]
            modelo = st.session_state.setdefault("_perfil_res", {}).setdefault(bid, {})
            
            # A chave do editor muda quando o conjunto de códigos muda
            ed_key = f"am_{bid}__{len(cods)}|" + "|".join(cods)
            
            # Chave para guardar o dataframe estático na sessão
            df_state_key = f"_base_df_{bid}"

            # SÓ recriamos o dataframe base se a lista de códigos de barras (RBC) 
            # tiver mudado ou se for a primeira vez renderizando. 
            # Isso evita a sobreposição de dados que apaga a digitação rápida.
            if st.session_state.get(f"_prev_key_{bid}") != ed_key or df_state_key not in st.session_state:
                chaves, r1s, r2s = [], [], []
                for i, c in enumerate(cods):
                    chave = c if c else f"__pos_{i}"
                    chaves.append(chave)
                    r1, r2 = modelo.get(chave, ("", ""))
                    r1s.append(r1)
                    r2s.append(r2)
                
                dados = pd.DataFrame({"Código de barras": cods,
                                      "Resultado 1": r1s, "Resultado 2": r2s})
                
                st.session_state[df_state_key] = dados
                st.session_state[f"_prev_key_{bid}"] = ed_key
            else:
                # Se não houve mudança nos códigos, usamos o dataframe já instanciado
                dados = st.session_state[df_state_key]

            entrada = st.data_editor(
                dados, num_rows="fixed", use_container_width=True, key=ed_key,
                disabled=["Código de barras"],
                column_config={
                    "Código de barras": st.column_config.TextColumn("Código de barras"),
                    "Resultado 1": st.column_config.TextColumn("Resultado 1"),
                    "Resultado 2": st.column_config.TextColumn("Resultado 2"),
                },
            )
            
            # Continua salvando no 'modelo' em background para não perder 
            # os resultados caso você adicione um novo código de barras lá no topo.
            for i in range(len(entrada)):
                chave = cods[i] if cods[i] else f"__pos_{i}"
                modelo[chave] = (entrada.iloc[i]["Resultado 1"],
                                 entrada.iloc[i]["Resultado 2"])
                
            st.caption(f"🔗 Códigos de barras replicados do teste **{nome_master}** — "
                       f"edite-os no bloco do **{nome_master}** para atualizar todos "
                       "ao mesmo tempo.")
        else:
            seed = pd.DataFrame({"Código de barras": ["", "", ""],
                                 "Resultado 1": ["", "", ""],
                                 "Resultado 2": ["", "", ""]})
            entrada = st.data_editor(
                seed, num_rows="dynamic", use_container_width=True, key=f"am_{bid}",
                column_config={
                    "Código de barras": st.column_config.TextColumn("Código de barras"),
                    "Resultado 1": st.column_config.TextColumn("Resultado 1"),
                    "Resultado 2": st.column_config.TextColumn("Resultado 2"),
                },
            )
            if eh_master and perfil_ativo:
                codigos_master = entrada["Código de barras"].tolist()
                nome_master = teste_sel
                st.caption(f"🔗 Perfil ativo: os códigos de barras deste teste "
                           f"(**{nome_master}**) são replicados automaticamente para os "
                           "demais testes do perfil.")

        blocos_dados.append((teste_sel, etm, ir_txt, lo, hi, zc_lo, zc_hi, entrada))

if st.button("➕ Adicionar teste"):
    st.session_state.imp_blocos.append(st.session_state.imp_next)
    st.session_state.imp_next += 1
    st.rerun()

# ---- Processar análise ---------------------------------------------------- #
# A análise (seções 4 e 5) só roda depois de clicar em "Processar análise".
# Uma assinatura leve das entradas indica se o que está na tela ainda corresponde
# ao que foi processado; enquanto o usuário digita, a assinatura muda e nada é
# recalculado (evita travamentos ao preencher vários testes).
_assinatura = "\n".join(
    [str(equip_sel), str(operador), str(data_problema)]
    + [f"{t[0]}::{t[7].to_csv(index=False)}" for t in blocos_dados]
)

st.markdown("")
if st.button("🔎 Processar análise", type="primary"):
    st.session_state["imp_proc_sig"] = _assinatura

if st.session_state.get("imp_proc_sig") != _assinatura:
    st.info("Preencha os resultados e clique em **🔎 Processar análise** para gerar as "
            "seções **4 · Resultado** e **5 · Exportar**. Enquanto você digita, nada é "
            "processado — clique novamente sempre que alterar algum dado.")
    st.stop()

# ---- Cálculo (todos os blocos) -------------------------------------------- #
partes, avisos = [], []
for teste_sel, etm, ir_txt, lo, hi, zc_lo, zc_hi, entrada in blocos_dados:
    res, q = analisar_bloco(entrada, teste_sel, etm, ir_txt, lo, hi, zc_lo, zc_hi)
    if res is None:
        if q > 0:   # bloco vazio não gera aviso; só o parcialmente preenchido (1-2)
            avisos.append((teste_sel, q))
    else:
        partes.append(res)

for teste_sel, q in avisos:
    st.warning(f"⚠️ **{teste_sel}**: informe no mínimo 3 amostras válidas (há {q}). "
               "Este teste não entrou na análise.")

if not partes:
    st.info("ℹ️ Informe pelo menos um teste com **3 amostras válidas** (código de barras + "
            "Resultado 1 e 2 numéricos, R1 ≠ 0) para ver a análise de impacto.")
    st.stop()

todos = pd.concat(partes, ignore_index=True)

if not operador.strip():
    st.warning("✏️ Preencha o **Nome do operador responsável** (seção 1) para gerar a análise.")
    st.stop()

# ---- 4 · Resultado da análise --------------------------------------------- #
st.markdown("### 4 · Resultado da análise de impacto")
st.caption(f"Equipamento avaliado: **{equip_sel}** · {todos['Teste'].nunique()} teste(s).")

n = len(todos)
n_etm = int(todos["Excede ETM"].sum())
n_interp = int(todos["Mudou interpretação"].sum())
n_impacto = int((todos["Impacto"] != "Sem impacto").sum())

m1, m2, m3, m4 = st.columns(4)
m1.metric("Amostras analisadas", f"{n}")
m2.metric("Excedem o ETM", f"{n_etm}")
m3.metric("Mudam de interpretação", f"{n_interp}")
m4.metric("Com impacto (combinado)", f"{n_impacto}")

if n_impacto:
    st.error(f"⚠️ {n_impacto} de {n} amostra(s) **com impacto** no equipamento **{equip_sel}** "
             "(excedem o ETM e/ou mudam de interpretação). Avalie antes de liberar.")
else:
    st.success(f"✅ Nenhuma das {n} amostras apresentou impacto em **{equip_sel}**.")

# Resumo por teste
resumo = (todos.assign(_imp=(todos["Impacto"] != "Sem impacto").astype(int),
                       _etm=todos["Excede ETM"].astype(int),
                       _int=todos["Mudou interpretação"].astype(int))
          .groupby("Teste")
          .agg(Amostras=("Impacto", "size"), Excedem_ETM=("_etm", "sum"),
               Mudam_interp=("_int", "sum"), Com_impacto=("_imp", "sum"))
          .reset_index())
st.markdown("**Resumo por teste**")
st.dataframe(resumo, use_container_width=True)

# Detalhe por amostra
tab = todos[["Teste", "Código de barras", "R1", "R2", "Erro total %", "Excede ETM",
             "Interpretação R1", "Interpretação R2", "Mudou interpretação", "Impacto"]].rename(
    columns={"R1": "Resultado 1", "R2": "Resultado 2"})


def _hl(v):
    # verde claro (aprovado) quando sem impacto; vermelho claro nas discordâncias
    return ("background-color:#E7F6EC; color:#0F5132" if v == "Sem impacto"
            else "background-color:#FFE3E3; color:#9B1C1C; font-weight:700")


st.markdown("**Detalhe por amostra**")
st.dataframe(tab.style.format({"Erro total %": "{:.2f}", "Resultado 1": "{:.3f}",
                               "Resultado 2": "{:.3f}"}).map(_hl, subset=["Impacto"]),
             use_container_width=True)

# ---- 5 · Exportar --------------------------------------------------------- #
st.markdown("### 5 · Exportar")
export = todos[["Teste", "ETM (%)", "IR", "Código de barras", "R1", "R2", "Erro total %",
                "Excede ETM", "Interpretação R1", "Interpretação R2",
                "Mudou interpretação", "Impacto"]].rename(
    columns={"R1": "Resultado 1", "R2": "Resultado 2"}).copy()
export.insert(0, "Equipamento", equip_sel)
export.insert(0, "Operador", operador)
export["Erro total %"] = export["Erro total %"].round(2)

# Nome padrão dos arquivos: "Análise de Impacto [equipamento] - [data DD-MM-AAAA]".
# A data é a do problema (seção 2); se vazia, usa a data de geração (hoje, Brasília).
_data_arq = (data_problema.strftime("%d-%m-%Y") if data_problema
             else datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%d-%m-%Y"))
_nome_arq = re.sub(r'[\\/:*?"<>|]+', "-",
                   f"Análise de Impacto {equip_sel} - {_data_arq}").strip()

d1, d2, d3 = st.columns(3)
with d1:
    st.download_button("⬇️ Baixar (Excel)",
                       data=to_excel(export, cols_2dec=["Erro total %", "Resultado 1", "Resultado 2"]),
                       file_name=f"{_nome_arq}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
with d2:
    csv_bytes = export.to_csv(index=False, sep=";", decimal=",", encoding="utf-8-sig").encode("utf-8-sig")
    st.download_button("⬇️ Baixar (CSV)", data=csv_bytes,
                       file_name=f"{_nome_arq}.csv", mime="text/csv")
with d3:
    data_prob_txt = data_problema.strftime("%d/%m/%Y") if data_problema else ""
    st.download_button("⬇️ Baixar (PDF)",
                       data=gerar_pdf(tab, equip_sel, operador, data_prob_txt),
                       file_name=f"{_nome_arq}.pdf", mime="application/pdf")

# -*- coding: utf-8 -*-
"""
Análise de Repetições (Duplicatas) — DataSift
==============================================

Ferramenta de controle da qualidade analítica baseada em resultados repetidos
(R1 e R2) de amostras de pacientes. O objetivo é detectar problemas de
equipamento/reagente avaliando o quanto a segunda medição (R2) concorda com a
primeira (R1).

Blocos de análise:
  3) Avaliação das repetições: marca cada amostra como OK/Suspeita combinando o
     erro total |(R1−R2)/R1| acima de um limite e a mudança de interpretação
     (intervalo de referência da própria planilha ou informado manualmente).
  4) Gráficos de apoio: Bland-Altman e média móvel das diferenças (deriva).

Cada amostra é identificada pelo código de barras, para rastrear qual paciente
ficou suspeito. A média móvel das diferenças pode ser ordenada pela data/hora
do resultado.

Entrada dos dados: (A) uma planilha já com R1 e R2 na mesma linha; ou (B) dois
relatórios (original e repetição) que o script junta automaticamente por
código de barras + teste (equivalente ao PROCV do Excel).

Este arquivo é uma PÁGINA do app DataSift (pasta ``pages/``), mas também roda
de forma independente com ``streamlit run pages/1_Analise_de_Repeticoes.py``.
"""

import io
import os
import re
import zipfile
import tempfile

import numpy as np
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt

# --------------------------------------------------------------------------- #
# Configuração de página / identidade visual (mesma paleta do DataSift)
# --------------------------------------------------------------------------- #
COLOR_PRIMARY = "#073B4C"
COLOR_SECONDARY = "#00E5FF"
COLOR_TERTIARY = "#118AB2"
COLOR_BG = "#F8F9FA"

try:
    st.set_page_config(page_title="DataSift · Repetições", page_icon="🔁", layout="wide")
except Exception:
    # Em app multipágina o set_page_config pode já ter sido definido; ignore.
    pass

st.markdown(
    f"""
    <style>
        .stApp {{ background-color: {COLOR_BG} !important; }}
        h1, h2, h3, h4 {{ color: {COLOR_PRIMARY} !important; font-weight: 800 !important; }}
        p, span, div[data-testid="stMarkdownContainer"], label {{ color: #212529 !important; }}
        div[data-testid="stMetricValue"] {{ color: {COLOR_PRIMARY} !important; }}
    </style>
    """,
    unsafe_allow_html=True,
)


# --------------------------------------------------------------------------- #
# 1. Normalização numérica (decimais com "," ou ".", milhar, unidades, etc.)
# --------------------------------------------------------------------------- #
def normalizar_serie_numerica(serie: pd.Series) -> pd.Series:
    """
    Converte uma coluna de texto/misto para float, lidando com:
      - vírgula OU ponto como separador decimal ("12,5" e "12.5");
      - separador de milhar ("1.234,56" -> 1234.56 e "1,234.56" -> 1234.56);
      - unidades/símbolos junto ao número ("12,5 mg/dL", "> 200") -> extrai 12.5 / 200;
      - células vazias, textos ("indetectável", "N/A") -> NaN.

    Regra do separador decimal: quando há "," e "." na mesma célula, o separador
    decimal é o que aparece POR ÚLTIMO; o outro é tratado como milhar. Quando há
    apenas ",", ela é tratada como decimal (padrão brasileiro).
    """
    def _conv(x):
        if pd.isna(x):
            return np.nan
        s = str(x).strip()
        if s == "" or s.lower() in ("nan", "none", "na", "n/a", "-", "--", "."):
            return np.nan
        # remove espaços (inclusive não separável) e mantém só dígitos/sinais/separadores
        s = s.replace("\xa0", "").replace(" ", "")
        s = re.sub(r"[^0-9,.\-+]", "", s)
        if s in ("", "+", "-", ".", ","):
            return np.nan
        has_dot, has_comma = "." in s, "," in s
        if has_dot and has_comma:
            if s.rfind(",") > s.rfind("."):      # vírgula é o decimal
                s = s.replace(".", "").replace(",", ".")
            else:                                # ponto é o decimal
                s = s.replace(",", "")
        elif has_comma:                          # só vírgula -> decimal
            s = s.replace(",", ".")
        # se houver mais de um ponto (milhar tipo "1.234.567"), remove todos menos o último
        if s.count(".") > 1:
            partes = s.split(".")
            s = "".join(partes[:-1]) + "." + partes[-1]
        try:
            return float(s)
        except ValueError:
            return np.nan

    return serie.apply(_conv)


def parse_limite(txt: str):
    """Converte um limite digitado (aceita vírgula) para float; vazio -> None."""
    if txt is None:
        return None
    v = normalizar_serie_numerica(pd.Series([txt])).iloc[0]
    return None if pd.isna(v) else float(v)


# --------------------------------------------------------------------------- #
# 2. Leitura robusta da planilha (csv / xlsx / xls / zip)
# --------------------------------------------------------------------------- #
def _ler_csv(path):
    """Tenta o padrão brasileiro (;, decimal ,) e cai para (, decimal .)."""
    try:
        return pd.read_csv(path, sep=";", decimal=",", encoding="latin-1", engine="python")
    except Exception:
        return pd.read_csv(path, sep=",", decimal=".", encoding="utf-8", engine="python")


@st.cache_data(show_spinner="Lendo planilha...")
def carregar_planilha(conteudo: bytes, nome: str) -> pd.DataFrame:
    """Recebe os bytes do arquivo enviado e devolve um DataFrame."""
    nome = nome.lower()
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(nome)[1]) as tmp:
        tmp.write(conteudo)
        tmp_path = tmp.name
    try:
        if nome.endswith(".zip"):
            with zipfile.ZipFile(tmp_path) as z:
                validos = [f for f in z.namelist()
                           if not f.startswith("__MACOSX/")
                           and f.lower().endswith((".csv", ".xlsx", ".xls"))]
                if not validos:
                    raise ValueError("O ZIP não contém CSV ou Excel válidos.")
                with tempfile.NamedTemporaryFile(delete=False,
                                                 suffix=os.path.splitext(validos[0])[1]) as inner:
                    inner.write(z.read(validos[0]))
                    inner_path = inner.name
                df = (_ler_csv(inner_path) if validos[0].lower().endswith(".csv")
                      else pd.read_excel(inner_path, engine="openpyxl"))
                os.remove(inner_path)
        elif nome.endswith(".csv"):
            df = _ler_csv(tmp_path)
        else:
            df = pd.read_excel(tmp_path, engine="openpyxl")
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    return df


def montar_datahora(df: pd.DataFrame, col_data: str | None, col_hora: str | None):
    """Combina as colunas de data e hora num datetime (dd/mm/aaaa, dayfirst)."""
    if not col_data or col_data not in df.columns:
        return None
    d = df[col_data].astype(str).str.strip()
    if col_hora and col_hora in df.columns:
        combo = d + " " + df[col_hora].astype(str).str.strip()
    else:
        combo = d
    return pd.to_datetime(combo, errors="coerce", dayfirst=True)


# --------------------------------------------------------------------------- #
# 2b. Junção de dois relatórios por código de barras + teste (tipo PROCV)
# --------------------------------------------------------------------------- #
def _chave_barcode(serie: pd.Series) -> pd.Series:
    """Normaliza o código de barras para casar entre planilhas (tira espaços e '.0')."""
    def f(x):
        if pd.isna(x):
            return ""
        s = str(x).strip().replace("\xa0", "")
        if re.fullmatch(r"\d+\.0", s):   # código lido como float (123.0) -> 123
            s = s[:-2]
        return s
    return serie.apply(f)


def _chave_teste(serie: pd.Series) -> pd.Series:
    """Normaliza o nome do teste para casar (minúsculas, espaços colapsados)."""
    def f(x):
        if pd.isna(x):
            return ""
        return re.sub(r"\s+", " ", str(x).strip()).casefold()
    return serie.apply(f)


def _guess_idx(cols, termos, default=0):
    """Índice da 1ª coluna cujo nome contém um dos termos (pré-seleção dos selects)."""
    for i, c in enumerate(cols):
        cl = str(c).lower()
        if any(t in cl for t in termos):
            return i
    return default


def juntar_relatorios(df1, df2, id1, ts1, res1, id2, ts2, res2, data1=None, hora1=None,
                      extras1=None, extras2=None):
    """
    Junta o relatório original (df1) com o das repetições (df2) casando pela chave
    composta código de barras + teste — equivale ao PROCV do Excel, mas usando duas
    colunas como chave (um mesmo código de barras pode ter mais de um teste).

    Devolve:
      - ``matched``: pares completos, com colunas Código de barras, Teste, R1, R2
        (e _data/_hora, se informadas no original);
      - ``status``: todas as amostras (casadas e não casadas) com a coluna Status;
      - ``stats``: contadores da junção.
    """
    a = pd.DataFrame({
        "_bc": _chave_barcode(df1[id1]).values,
        "_ts": _chave_teste(df1[ts1]).values,
        "bc1": df1[id1].astype(str).values,
        "ts1": df1[ts1].astype(str).values,
        "R1": df1[res1].values,
    })
    if data1:
        a["_data"] = df1[data1].values
    if hora1:
        a["_hora"] = df1[hora1].values
    if extras1:
        for _n, _c in extras1.items():
            a[_n] = df1[_c].values
    b = pd.DataFrame({
        "_bc": _chave_barcode(df2[id2]).values,
        "_ts": _chave_teste(df2[ts2]).values,
        "bc2": df2[id2].astype(str).values,
        "ts2": df2[ts2].astype(str).values,
        "R2": df2[res2].values,
    })
    if extras2:
        for _n, _c in extras2.items():
            b[_n] = df2[_c].values
    a = a[a["_bc"] != ""]
    b = b[b["_bc"] != ""]
    dup1 = int(a.duplicated(subset=["_bc", "_ts"]).sum())
    dup2 = int(b.duplicated(subset=["_bc", "_ts"]).sum())
    a = a.drop_duplicates(subset=["_bc", "_ts"], keep="first")
    b = b.drop_duplicates(subset=["_bc", "_ts"], keep="first")

    m = a.merge(b, on=["_bc", "_ts"], how="outer", indicator=True)
    m["Código de barras"] = m["bc1"].fillna(m["bc2"])
    m["Teste"] = m["ts1"].fillna(m["ts2"])
    status_map = {"both": "Par completo",
                  "left_only": "Só no original (sem repetição)",
                  "right_only": "Só na repetição (sem original)"}
    m["Status"] = m["_merge"].map(status_map)

    stats = {
        "n1": int(len(a)), "n2": int(len(b)),
        "n_match": int((m["_merge"] == "both").sum()),
        "n_so_orig": int((m["_merge"] == "left_only").sum()),
        "n_so_rep": int((m["_merge"] == "right_only").sum()),
        "dup1": dup1, "dup2": dup2,
    }
    extra = [c for c in (["_data", "_hora"] + list((extras1 or {}).keys())
                         + list((extras2 or {}).keys())) if c in m.columns]
    matched = (m[m["_merge"] == "both"][["Código de barras", "Teste", "R1", "R2"] + extra]
               .reset_index(drop=True))
    status = m[["Código de barras", "Teste", "R1", "R2", "Status"]].reset_index(drop=True)
    return matched, status, stats


# --------------------------------------------------------------------------- #
# 3. Cálculos estatísticos das duplicatas
# --------------------------------------------------------------------------- #
def calcular_metricas(df: pd.DataFrame, col_r1: str, col_r2: str,
                      col_id: str | None = None, datahora=None, extras=None, z: float = 1.96):
    """
    Recebe o DataFrame e devolve:
      - tabela por par (ID, R1, R2, média, diferença, erro relativo %, DataHora)
      - dicionário com as métricas agregadas.
    A coluna ``ID`` guarda o código de barras (ou o nº da linha, se não informado).
    """
    n = len(df)
    if col_id and col_id in df.columns:
        ids = df[col_id].astype(str).values
    else:
        ids = np.array([f"linha {i + 1}" for i in range(n)])

    dados = {
        "ID": ids,
        "R1": normalizar_serie_numerica(df[col_r1]).values,
        "R2": normalizar_serie_numerica(df[col_r2]).values,
    }
    if datahora is not None:
        dados["DataHora"] = pd.to_datetime(datahora).values
    base = pd.DataFrame(dados)
    # Colunas adicionais (equipamento, idade, sexo, etc.) alinhadas por posição:
    if extras:
        for _nome, _val in extras.items():
            base[_nome] = np.asarray(_val)

    n_total = len(base)
    base = base.dropna(subset=["R1", "R2"])
    base = base[base["R1"] != 0].reset_index(drop=True)  # evita divisão por zero em (R1-R2)/R1
    n_validos = len(base)

    base["Media_par"] = (base["R1"] + base["R2"]) / 2.0
    base["Diferenca"] = base["R1"] - base["R2"]              # d = R1 - R2
    base["Dif_abs"] = base["Diferenca"].abs()
    # DP e CV de cada par (n=2): DP = |R1-R2|/sqrt(2)
    base["DP_par"] = base["Dif_abs"] / np.sqrt(2)
    base["CV_par_%"] = np.where(base["Media_par"] != 0,
                                base["DP_par"] / base["Media_par"] * 100, np.nan)
    # Erro total analítico (fórmula do usuário): (R1 - R2) / R1  em %
    base["ETA_%"] = base["Diferenca"] / base["R1"] * 100

    d = base["Diferenca"].to_numpy()
    media_global = float(base["Media_par"].mean()) if n_validos else np.nan
    # DP de repetibilidade (a partir das diferenças): Sr = sqrt( sum(d^2) / (2n) )
    dp_repet = float(np.sqrt(np.sum(d ** 2) / (2 * n_validos))) if n_validos else np.nan
    cv_analitico = (dp_repet / media_global * 100) if media_global else np.nan
    vies_medio = float(base["Diferenca"].mean()) if n_validos else np.nan   # bias
    vies_medio_pct = (vies_medio / media_global * 100) if media_global else np.nan
    erro_aleatorio = z * dp_repet                                            # DP x 1,96
    # Erro total analítico "clássico" (modelo Westgard): |bias%| + z * CV%
    eta_westgard = abs(vies_medio_pct) + z * cv_analitico if n_validos else np.nan

    resumo = {
        "n_total": n_total,
        "n_validos": n_validos,
        "media_global": media_global,
        "dp_repet": dp_repet,
        "cv_analitico": cv_analitico,
        "vies_medio": vies_medio,
        "vies_medio_pct": vies_medio_pct,
        "erro_aleatorio": erro_aleatorio,
        "eta_medio_pct": float(base["ETA_%"].mean()) if n_validos else np.nan,
        "eta_westgard": eta_westgard,
        "z": z,
    }
    return base, resumo


# --------------------------------------------------------------------------- #
# 4. Classificação por intervalo de referência / limite de decisão médica
# --------------------------------------------------------------------------- #
def parse_ref_range(txt):
    """
    Extrai (limite_inferior, limite_superior) de um texto de intervalo de
    referência. Aceita formatos como '136.00 - 145.00', '0,5 - 1,2', '< 200',
    '> 40', '<= 5'. Devolve None no limite que não existir; (None, None) quando
    não consegue interpretar (ex.: 'Negativo', vazio).
    """
    if pd.isna(txt):
        return (None, None)
    s = str(txt).strip()
    if s == "":
        return (None, None)
    low = s.lower()
    # números sem sinal (o sinal de intervalo '-' não deve virar negativo):
    nums = re.findall(r"\d+(?:[.,]\d+)?", s)
    vals = [float(n.replace(",", ".")) for n in nums]
    if not vals:
        return (None, None)
    if ("<" in s or "≤" in s or "menor" in low or "até" in low or "up to" in low):
        return (None, vals[-1])
    if (">" in s or "≥" in s or "maior" in low or "acima" in low):
        return (vals[0], None)
    if len(vals) >= 2:
        return (min(vals[0], vals[1]), max(vals[0], vals[1]))
    return (None, None)   # um único número sem sinal -> ambíguo, não classifica


def classificar_ref(valor, limite_inf, limite_sup):
    """
    Classifica um valor em Baixo / Normal / Alto a partir do intervalo de
    referência. Limites são inclusivos (Normal = inf <= valor <= sup). Qualquer
    limite pode ficar vazio (None/NaN) para representar "sem limite" daquele lado;
    se ambos estiverem vazios, devolve '—' (sem intervalo definido).
    """
    if pd.isna(valor):
        return "—"
    inf_ok = limite_inf is not None and pd.notna(limite_inf)
    sup_ok = limite_sup is not None and pd.notna(limite_sup)
    if not inf_ok and not sup_ok:
        return "—"
    if inf_ok and valor < limite_inf:
        return "Baixo"
    if sup_ok and valor > limite_sup:
        return "Alto"
    return "Normal"


# --------------------------------------------------------------------------- #
# 5. Exportação
# --------------------------------------------------------------------------- #
def to_excel(df: pd.DataFrame, cols_2dec=None) -> bytes:
    """
    Exporta em .xlsx com todo o conteúdo centralizado, largura das colunas
    ajustada ao conteúdo (autofit) e 2 casas decimais nas colunas indicadas.
    """
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    cols_2dec = set(cols_2dec or [])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Repeticoes")
        ws = writer.sheets["Repeticoes"]
        centro = Alignment(horizontal="center", vertical="center")
        n = len(df)
        for i, col in enumerate(df.columns, start=1):
            letra = get_column_letter(i)
            # autofit: largura = maior conteúdo/cabeçalho (limitada a 60)
            largura = max([len(str(col))] + [len(str(v)) for v in df[col].values])
            ws.column_dimensions[letra].width = min(largura + 2, 60)
            for r in range(1, n + 2):                       # centraliza cabeçalho + dados
                ws.cell(row=r, column=i).alignment = centro
            if col in cols_2dec:                            # exibe 2 casas decimais
                for r in range(2, n + 2):
                    ws.cell(row=r, column=i).number_format = "0.00"
    return output.getvalue()


# =========================================================================== #
#                                INTERFACE
# =========================================================================== #
st.markdown("## 🔁 Análise de Repetições (Duplicatas)")
st.caption(
    "Avalia a concordância entre um primeiro resultado (R1) e sua repetição (R2) "
    "para detectar problemas de equipamento/reagente, e verifica se a interpretação "
    "clínica mudou entre as duas medições"
)

# ---- 1 · Origem dos dados ------------------------------------------------- #
with st.container(border=True):
    st.markdown("### 1 · Origem dos dados")
    modo = st.radio(
        "Como você vai fornecer os dados?",
        ["Uma planilha (R1 e R2 já na mesma linha)",
         "Dois relatórios (original + repetição) — juntar por código de barras + teste"],
    )
dois_relatorios = modo.startswith("Dois")

# Variáveis que os dois modos preenchem antes da análise:
col_r1 = col_r2 = col_id = col_analito = col_data = col_hora = None
col_equip1 = col_equip2 = col_r1ant = col_idade = col_sexo = col_ref = col_valid1 = None
df = None

if not dois_relatorios:
    # ===================== MODO A · uma planilha ===========================
    with st.container(border=True):
        st.markdown("#### Envie a planilha de resultados")
        arquivo = st.file_uploader(
            "Arquivo com as colunas R1, R2 e o código de barras (uma linha por amostra)",
            type=["csv", "xlsx", "xls", "zip"], key="upload_unico",
        )
    if arquivo is None:
        st.info(
            "📄 Envie um arquivo para começar. Ele deve conter, no mínimo, três colunas: "
            "o primeiro resultado (R1), a repetição (R2) e o código de barras que identifica "
            "a amostra. Colunas de analito, data e hora são opcionais."
        )
        st.stop()
    df = carregar_planilha(arquivo.getvalue(), arquivo.name)
    if df is None or df.empty:
        st.error("Não foi possível ler a planilha ou ela está vazia.")
        st.stop()
    st.success(f"Planilha carregada: {len(df)} linhas × {len(df.columns)} colunas.")

    with st.container(border=True):
        st.markdown("### 2 · Indique as colunas")
        colunas = list(df.columns)
        opc = ["(nenhuma)"] + colunas
        c1, c2, c3 = st.columns(3)
        with c1:
            col_r1 = st.selectbox("Coluna do **R1** (1º resultado)", colunas, index=0)
        with c2:
            idx2 = 1 if len(colunas) > 1 else 0
            col_r2 = st.selectbox("Coluna do **R2** (repetição)", colunas, index=idx2)
        with c3:
            _cid = st.selectbox("Coluna do **código de barras** (identificador)",
                                ["(usar nº da linha)"] + colunas, index=0)
            col_id = None if _cid == "(usar nº da linha)" else _cid
        c4, c5, c6 = st.columns(3)
        with c4:
            col_analito = st.selectbox("Coluna de analito/exame (opcional)", opc, index=0)
        with c5:
            _cd = st.selectbox("Coluna de **data** (opcional)", opc, index=0)
            col_data = None if _cd == "(nenhuma)" else _cd
        with c6:
            _ch = st.selectbox("Coluna de **hora** (opcional)", opc, index=0)
            col_hora = None if _ch == "(nenhuma)" else _ch

        st.markdown("**Colunas opcionais** — aparecem na tabela do Bloco 3; a coluna de "
                    "intervalo de referência habilita a opção *“usar o do sistema”*.")
        o1, o2, o3 = st.columns(3)
        with o1:
            _e1 = st.selectbox("Equipamento do R1", opc, index=0, key="equip1")
            col_equip1 = None if _e1 == "(nenhuma)" else _e1
        with o2:
            _e2 = st.selectbox("Equipamento do R2", opc, index=0, key="equip2")
            col_equip2 = None if _e2 == "(nenhuma)" else _e2
        with o3:
            _ra = st.selectbox("Resultado anterior do R1", opc, index=0, key="r1ant")
            col_r1ant = None if _ra == "(nenhuma)" else _ra
        o4, o5, o6 = st.columns(3)
        with o4:
            _id = st.selectbox("Idade do paciente", opc, index=0, key="idade")
            col_idade = None if _id == "(nenhuma)" else _id
        with o5:
            _sx = st.selectbox("Sexo do paciente", opc, index=0, key="sexo")
            col_sexo = None if _sx == "(nenhuma)" else _sx
        with o6:
            _rf = st.selectbox("Intervalo de referência (ex.: Ref. ranges)", opc, index=0, key="refcol")
            col_ref = None if _rf == "(nenhuma)" else _rf
        o7, _o8 = st.columns(2)
        with o7:
            _uv = st.selectbox("Usuário de validação do 1º resultado", opc, index=0, key="valid1")
            col_valid1 = None if _uv == "(nenhuma)" else _uv

        z_opt = st.selectbox("Nível de confiança (Z)",
                             ["95% bilateral (Z = 1,96)", "95% unilateral (Z = 1,65)"], index=0)
        z = 1.96 if "1,96" in z_opt else 1.65

    if col_r1 == col_r2:
        st.warning("R1 e R2 estão apontando para a mesma coluna. Selecione colunas diferentes.")
        st.stop()

else:
    # ============= MODO B · dois relatórios (junção tipo PROCV) =============
    with st.container(border=True):
        st.markdown("#### Envie os dois relatórios")
        u1, u2 = st.columns(2)
        with u1:
            arq1 = st.file_uploader("1️⃣ Relatório **original** (vira R1)",
                                    type=["csv", "xlsx", "xls", "zip"], key="upload_orig")
        with u2:
            arq2 = st.file_uploader("2️⃣ Relatório da **repetição** (vira R2)",
                                    type=["csv", "xlsx", "xls", "zip"], key="upload_rep")
    if arq1 is None or arq2 is None:
        st.info(
            "📄 Envie os **dois** relatórios. O script casa as amostras por "
            "**código de barras + teste** (como um PROCV) e monta os pares R1/R2 — "
            "usar o teste na chave é importante porque um mesmo código de barras pode "
            "ter mais de um teste."
        )
        st.stop()
    df1 = carregar_planilha(arq1.getvalue(), arq1.name)
    df2 = carregar_planilha(arq2.getvalue(), arq2.name)
    if df1 is None or df1.empty or df2 is None or df2.empty:
        st.error("Não foi possível ler um dos relatórios (ou algum está vazio).")
        st.stop()
    st.success(f"Original: {len(df1)}×{len(df1.columns)}  ·  Repetição: {len(df2)}×{len(df2.columns)}.")

    with st.container(border=True):
        st.markdown("### 2 · Indique as colunas de cada relatório")
        st.caption("Código de barras + teste são a **chave** que casa as amostras; o "
                   "resultado de cada relatório vira R1 (original) e R2 (repetição).")
        cols1, cols2 = list(df1.columns), list(df2.columns)

        st.markdown("**Relatório original (R1)**")
        a1, a2, a3 = st.columns(3)
        with a1:
            id1 = st.selectbox("Código de barras", cols1,
                               index=_guess_idx(cols1, ["barra", "cod", "amostra"]), key="id1")
        with a2:
            ts1 = st.selectbox("Teste/exame", cols1,
                               index=_guess_idx(cols1, ["teste", "exame", "analito", "prova"]), key="ts1")
        with a3:
            res1 = st.selectbox("Resultado → R1", cols1,
                                index=_guess_idx(cols1, ["result", "valor", "dosagem"]), key="res1")
        opc1 = ["(nenhuma)"] + cols1
        a4, a5, a6 = st.columns(3)
        with a4:
            _d1 = st.selectbox("Data do R1 (opcional)", opc1, index=0, key="data1")
            data1 = None if _d1 == "(nenhuma)" else _d1
        with a5:
            _h1 = st.selectbox("Hora do R1 (opcional)", opc1, index=0, key="hora1")
            hora1 = None if _h1 == "(nenhuma)" else _h1
        with a6:
            _rf1 = st.selectbox("Intervalo de referência (opcional)", opc1, index=0, key="ref1b")
            ref1 = None if _rf1 == "(nenhuma)" else _rf1
        a7, a8, a9 = st.columns(3)
        with a7:
            _eq1 = st.selectbox("Equipamento do R1 (opcional)", opc1, index=0, key="eq1b")
            eq1 = None if _eq1 == "(nenhuma)" else _eq1
        with a8:
            _ra1 = st.selectbox("Resultado anterior do R1 (opcional)", opc1, index=0, key="ra1b")
            ra1 = None if _ra1 == "(nenhuma)" else _ra1
        with a9:
            _ida1 = st.selectbox("Idade do paciente (opcional)", opc1, index=0, key="ida1b")
            ida1 = None if _ida1 == "(nenhuma)" else _ida1
        a10, a11 = st.columns(2)
        with a10:
            _sex1 = st.selectbox("Sexo do paciente (opcional)", opc1, index=0, key="sex1b")
            sex1 = None if _sex1 == "(nenhuma)" else _sex1
        with a11:
            _uv1 = st.selectbox("Usuário de validação do R1 (opcional)", opc1, index=0, key="valid1b")
            valid1 = None if _uv1 == "(nenhuma)" else _uv1

        st.markdown("**Relatório da repetição (R2)**")
        b1, b2, b3 = st.columns(3)
        with b1:
            id2 = st.selectbox("Código de barras", cols2,
                               index=_guess_idx(cols2, ["barra", "cod", "amostra"]), key="id2")
        with b2:
            ts2 = st.selectbox("Teste/exame", cols2,
                               index=_guess_idx(cols2, ["teste", "exame", "analito", "prova"]), key="ts2")
        with b3:
            res2 = st.selectbox("Resultado → R2", cols2,
                                index=_guess_idx(cols2, ["result", "valor", "dosagem"]), key="res2")
        b4, _b5 = st.columns(2)
        with b4:
            _eq2 = st.selectbox("Equipamento do R2 (opcional)", ["(nenhuma)"] + cols2, index=0, key="eq2b")
            eq2 = None if _eq2 == "(nenhuma)" else _eq2

        z_opt = st.selectbox("Nível de confiança (Z)",
                             ["95% bilateral (Z = 1,96)", "95% unilateral (Z = 1,65)"], index=0)
        z = 1.96 if "1,96" in z_opt else 1.65

    extras1_map = {}
    for _nm, _cl in [("Equip. R1", eq1), ("R1 anterior", ra1), ("Idade", ida1),
                     ("Sexo", sex1), ("RefRange", ref1), ("Usuário validação R1", valid1)]:
        if _cl:
            extras1_map[_nm] = _cl
    extras2_map = {"Equip. R2": eq2} if eq2 else {}
    matched, status_merge, stats = juntar_relatorios(
        df1, df2, id1, ts1, res1, id2, ts2, res2, data1=data1, hora1=hora1,
        extras1=extras1_map, extras2=extras2_map)

    st.markdown("#### Resultado da junção (PROCV por código de barras + teste)")
    j1, j2, j3, j4 = st.columns(4)
    j1.metric("Pares casados (R1+R2)", f"{stats['n_match']}")
    j2.metric("Só no original", f"{stats['n_so_orig']}",
              help="Amostras/testes do relatório original que não tiveram repetição.")
    j3.metric("Só na repetição", f"{stats['n_so_rep']}",
              help="Repetições sem correspondente no relatório original.")
    j4.metric("Chaves duplicadas removidas", f"{stats['dup1'] + stats['dup2']}",
              help="Código de barras + teste repetidos dentro de um mesmo relatório "
                   "(mantida a 1ª ocorrência).")

    if stats["n_match"] == 0:
        st.error("Nenhuma amostra casou por código de barras + teste. Confira se as "
                 "colunas de código de barras e de teste estão corretas nos dois relatórios.")
        st.stop()

    if stats["n_so_orig"] or stats["n_so_rep"]:
        with st.expander(f"🔎 Ver {stats['n_so_orig'] + stats['n_so_rep']} amostra(s) não casada(s)"):
            nc = status_merge[status_merge["Status"] != "Par completo"]
            st.dataframe(nc, use_container_width=True, height=240)
            st.download_button(
                "⬇️ Baixar não casadas (CSV)",
                data=nc.to_csv(index=False, sep=";", decimal=",",
                               encoding="utf-8-sig").encode("utf-8-sig"),
                file_name="amostras_nao_casadas.csv", mime="text/csv")

    df = matched
    col_r1, col_r2, col_id, col_analito = "R1", "R2", "Código de barras", "Teste"
    col_data = "_data" if "_data" in matched.columns else None
    col_hora = "_hora" if "_hora" in matched.columns else None
    col_equip1 = "Equip. R1" if "Equip. R1" in matched.columns else None
    col_equip2 = "Equip. R2" if "Equip. R2" in matched.columns else None
    col_r1ant = "R1 anterior" if "R1 anterior" in matched.columns else None
    col_idade = "Idade" if "Idade" in matched.columns else None
    col_sexo = "Sexo" if "Sexo" in matched.columns else None
    col_ref = "RefRange" if "RefRange" in matched.columns else None
    col_valid1 = "Usuário validação R1" if "Usuário validação R1" in matched.columns else None

# ---- Filtro opcional por analito/teste ------------------------------------ #
df_uso = df
if col_analito and col_analito != "(nenhuma)":
    valores = ["(todos)"] + sorted(df[col_analito].dropna().astype(str).unique().tolist())
    escolha = st.selectbox("Filtrar por analito/teste", valores, index=0)
    if escolha != "(todos)":
        df_uso = df[df[col_analito].astype(str) == escolha]

# ---- Colunas adicionais para exibir/avaliar (alinhadas por posição) ------- #
extras = {}
if col_data and col_data in df_uso.columns:
    _dd = pd.to_datetime(df_uso[col_data], errors="coerce", dayfirst=True)
    extras["Data R1"] = _dd.dt.strftime("%d/%m/%Y").fillna("").values
if col_hora and col_hora in df_uso.columns:
    extras["Hora R1"] = df_uso[col_hora].astype(str).replace({"NaT": "", "nan": ""}).values
for _nome, _col in [("Equip. R1", col_equip1), ("Equip. R2", col_equip2),
                    ("R1 anterior", col_r1ant), ("Idade", col_idade),
                    ("Sexo", col_sexo), ("RefRange", col_ref),
                    ("Usuário validação R1", col_valid1)]:
    if _col and _col in df_uso.columns:
        extras[_nome] = df_uso[_col].values

# ---- Cálculo -------------------------------------------------------------- #
datahora = montar_datahora(df_uso, col_data, col_hora)
base, resumo = calcular_metricas(df_uso, col_r1, col_r2, col_id=col_id,
                                 datahora=datahora, extras=extras, z=z)

if resumo["n_validos"] == 0:
    st.error(
        "Nenhum par R1/R2 válido após a normalização. Verifique se as colunas "
        "escolhidas contêm números (a normalização aceita vírgula ou ponto)."
    )
    st.stop()

descartados = resumo["n_total"] - resumo["n_validos"]
if descartados > 0:
    st.warning(
        f"{descartados} linha(s) descartada(s) por valor ausente/não numérico em "
        f"R1 ou R2, ou por R1 = 0 (indefinido em (R1−R2)/R1)."
    )

# ---- Bloco 3: avaliação das repetições ------------------------------------ #
st.markdown("### 3 · Avaliação das repetições")
st.caption(
    "Marca cada amostra como **OK** ou **Suspeita** combinando dois critérios: o "
    "**erro total** |(R1−R2)/R1| acima do limite aceitável e a **mudança de "
    "interpretação** (intervalo de referência / limite de decisão médica) entre R1 e R2."
)

# --- Critério 1: erro total máximo ---
lim_eta = st.number_input(
    "Limite de aceitação: Erro Total Máximo",
    min_value=0.0, value=10.0, step=0.5,
    help="Erro total admissível para |(R1−R2)/R1|, em %. Pares acima disto são "
         "sinalizados como suspeitos. Defina conforme a especificação do analito "
         "(variação biológica, CLIA, RDC).",
)
base["Suspeito_erro"] = base["ETA_%"].abs() > lim_eta

# --- Critério 2: intervalo de referência / limite de decisão médica ---
st.markdown("**Intervalo de referência / limite de decisão médica**")
if "RefRange" in base.columns:
    opcoes_ref = [f"Usar o do sistema (coluna: {col_ref})", "Inserir manualmente"]
else:
    st.caption("💡 Para preencher automaticamente pelo sistema, selecione a coluna do "
               "intervalo de referência (ex.: *Ref. ranges*) na **seção 2 › Colunas opcionais**.")
    opcoes_ref = ["Inserir manualmente"]
origem_ref = st.radio("De onde vem o intervalo?", opcoes_ref, horizontal=True,
                      label_visibility="collapsed")

tem_ref = False
if origem_ref.startswith("Usar o do sistema"):
    _par = base["RefRange"].apply(parse_ref_range)
    base["_lo"] = [p[0] for p in _par]
    base["_hi"] = [p[1] for p in _par]
    n_falha = int(sum(1 for p in _par if p[0] is None and p[1] is None))
    tem_ref = n_falha < len(base)
    st.caption("Cada amostra é avaliada pelo **seu próprio** intervalo de referência "
               "(o da coluna já considera teste, idade e sexo do paciente).")
    if n_falha:
        st.warning(f"{n_falha} amostra(s) com intervalo não interpretável — ficam como "
                   "'—' e não contam como mudança de interpretação.")
else:
    ci1, ci2 = st.columns(2)
    with ci1:
        txt_inf = st.text_input("Limite inferior do normal (deixe vazio se não usar)", value="")
    with ci2:
        txt_sup = st.text_input("Limite superior do normal (deixe vazio se não usar)", value="")
    lim_inf, lim_sup = parse_limite(txt_inf), parse_limite(txt_sup)
    if lim_inf is not None and lim_sup is not None and lim_inf > lim_sup:
        st.warning("O limite inferior é maior que o superior. Verifique os valores.")
    base["_lo"] = lim_inf
    base["_hi"] = lim_sup
    tem_ref = (lim_inf is not None) or (lim_sup is not None)
    if tem_ref:
        faixa_txt = (f"{lim_inf if lim_inf is not None else '−∞'} a "
                     f"{lim_sup if lim_sup is not None else '+∞'}")
        st.caption(f"Faixa normal: **{faixa_txt}** (limites inclusivos, aplicada a todas as amostras).")
    else:
        st.info("Sem intervalo definido: a avaliação usa **apenas** o critério de erro total.")

# --- Classificação e situação combinada ---
if tem_ref:
    base["Interp_R1"] = base.apply(lambda r: classificar_ref(r["R1"], r["_lo"], r["_hi"]), axis=1)
    base["Interp_R2"] = base.apply(lambda r: classificar_ref(r["R2"], r["_lo"], r["_hi"]), axis=1)
    base["Mudou_interp"] = ((base["Interp_R1"] != base["Interp_R2"])
                            & (base["Interp_R1"] != "—") & (base["Interp_R2"] != "—"))
else:
    base["Interp_R1"] = "—"
    base["Interp_R2"] = "—"
    base["Mudou_interp"] = False

base["Situacao"] = np.where(base["Suspeito_erro"] | base["Mudou_interp"], "Suspeito", "OK")

def _motivo(row):
    e, m = row["Suspeito_erro"], row["Mudou_interp"]
    if e and m:
        return "Erro total + Mudança de interpretação"
    if e:
        return "Erro total"
    if m:
        return "Mudança de interpretação"
    return "—"
base["Motivo"] = base.apply(_motivo, axis=1)

n_mudou = int(base["Mudou_interp"].sum())
n_erro = int(base["Suspeito_erro"].sum())
n_comb = int((base["Situacao"] == "Suspeito").sum())

cA, cB, cC = st.columns(3)
cA.metric("Mudança de interpretação", f"{n_mudou}")
cB.metric("Suspeitos pelo erro total", f"{n_erro}")
cC.metric("Suspeitos (combinado)", f"{n_comb}",
          help=f"Erro total |(R1−R2)/R1| > {lim_eta:.1f}% OU mudança de interpretação entre R1 e R2.")

if n_comb:
    st.error(f"⚠️ {n_comb} amostra(s) suspeita(s) por pelo menos um critério "
             "(erro total e/ou mudança de interpretação). Veja a coluna **Situação**.")
else:
    st.success("Nenhuma amostra suspeita pelos critérios combinados.")

# --- Tabela consolidada (com as colunas adicionais informadas) ---
tab3 = base.rename(columns={
    "ID": "Código de barras", "Interp_R1": "Interpretação R1",
    "Interp_R2": "Interpretação R2", "ETA_%": "(R1−R2)/R1 %",
    "Situacao": "Situação",
})
cols_extra = [c for c in ["Data R1", "Hora R1", "Equip. R1", "Equip. R2",
                          "R1 anterior", "Idade", "Sexo", "Usuário validação R1"]
              if c in tab3.columns]
cols_interp = ["Interpretação R1", "Interpretação R2"] if tem_ref else []
col_ordem = (["Código de barras"] + cols_extra + ["R1", "R2"] + cols_interp
             + ["(R1−R2)/R1 %", "Motivo", "Situação"])
tab3 = tab3[col_ordem]

def _hl(v):
    return ("background-color:#FFE3E3; color:#9B1C1C; font-weight:700" if v == "Suspeito"
            else "background-color:#E7F6EC; color:#0F5132")
st.dataframe(tab3.style.format(precision=3).map(_hl, subset=["Situação"]),
             use_container_width=True, height=360)

if n_comb:
    with st.expander(f"🔎 Ver só os {n_comb} suspeito(s)"):
        st.dataframe(tab3[tab3["Situação"] == "Suspeito"].style.format(precision=3),
                     use_container_width=True)

if tem_ref:
    with st.expander("🔀 Matriz de transição (R1 → R2)"):
        st.caption("Quantas amostras foram de cada interpretação em R1 (linhas) para "
                   "cada interpretação em R2 (colunas). A diagonal são as que não mudaram.")
        st.dataframe(pd.crosstab(base["Interp_R1"], base["Interp_R2"],
                                 rownames=["R1"], colnames=["R2"]),
                     use_container_width=True)

# ---- Bloco 4: gráficos de apoio (deriva e concordância) ------------------- #
st.markdown("### 4 · Gráficos de apoio (deriva e concordância)")
md = resumo["vies_medio"]
sd = base["Diferenca"].std(ddof=1)
g1, g2 = st.columns(2)

with g1:
    # Bland-Altman: média do par (x) vs diferença (y)
    fig, ax = plt.subplots(figsize=(5, 3.6))
    ax.scatter(base["Media_par"], base["Diferenca"], s=14, alpha=0.6, color=COLOR_TERTIARY)
    ax.axhline(md, color=COLOR_PRIMARY, lw=1.5, label=f"Viés = {md:.3f}")
    if pd.notna(sd):
        ax.axhline(md + 1.96 * sd, color="#EF476F", ls="--", lw=1, label="±1,96 DP")
        ax.axhline(md - 1.96 * sd, color="#EF476F", ls="--", lw=1)
    ax.set_xlabel("Média do par (R1+R2)/2")
    ax.set_ylabel("Diferença (R1−R2)")
    ax.set_title("Bland-Altman")
    ax.legend(fontsize=7)
    ax.grid(alpha=0.2)
    st.pyplot(fig, clear_figure=True)

with g2:
    # Média móvel da diferença — ordenada por data/hora se disponível, senão por ordem
    janela = st.slider("Janela da média móvel das diferenças", 3, 50, 10)
    tem_dt = "DataHora" in base.columns and base["DataHora"].notna().any()
    if tem_dt:
        ordf = base.dropna(subset=["DataHora"]).sort_values("DataHora").reset_index(drop=True)
        x, xlabel = ordf["DataHora"], "Data/hora do resultado"
        sem_dt = int(base["DataHora"].isna().sum())
    else:
        ordf = base.reset_index(drop=True)
        x, xlabel, sem_dt = ordf.index, "Ordem da amostra", 0
    y = ordf["Diferenca"]
    mm = y.rolling(janela, min_periods=1).mean()
    fig2, ax2 = plt.subplots(figsize=(5, 3.6))
    ax2.plot(x, y.values, ".", ms=4, alpha=0.35, color="#999", label="Diferença")
    ax2.plot(x, mm.values, "-", lw=2, color=COLOR_SECONDARY, label=f"Média móvel ({janela})")
    ax2.axhline(0, color=COLOR_PRIMARY, lw=1)
    ax2.set_xlabel(xlabel)
    ax2.set_ylabel("Diferença (R1−R2)")
    ax2.set_title("Média móvel da diferença (deriva do sistema)")
    ax2.legend(fontsize=7)
    ax2.grid(alpha=0.2)
    if tem_dt:
        fig2.autofmt_xdate()
    st.pyplot(fig2, clear_figure=True)
    if tem_dt and sem_dt:
        st.caption(f"{sem_dt} amostra(s) sem data/hora válida não entraram neste gráfico.")
    elif not tem_dt and (col_data or col_hora):
        st.caption("Não foi possível interpretar a data/hora; gráfico ordenado pela ordem da amostra.")

# Tabela dos pontos fora dos limites de concordância (±1,96 DP) do Bland-Altman
if pd.notna(sd) and sd > 0:
    lim_ba = 1.96 * sd
    fora = base[(base["Diferenca"] - md).abs() > lim_ba]
    st.markdown(f"**Amostras fora dos limites de concordância do Bland-Altman "
                f"(viés {md:.3f} ± {lim_ba:.3f})**")
    if len(fora):
        tab_fora = fora.rename(columns={"ID": "Código de barras",
                                        "Diferenca": "R1−R2"})[["Código de barras", "R1", "R2", "R1−R2"]]
        st.dataframe(tab_fora.style.format(precision=3), use_container_width=True)
    else:
        st.success("Nenhuma amostra fora dos limites de ±1,96 DP.")

with st.expander("ℹ️ Como interpretar os gráficos"):
    st.markdown(
        """
**Gráfico de Bland-Altman** — mostra, para cada amostra, a **média do par**
`(R1+R2)/2` no eixo X e a **diferença** `R1−R2` no eixo Y. Serve para enxergar a
concordância entre a 1ª e a 2ª medição ao longo de toda a faixa de concentração.

- A **linha cheia central** é o **viés médio**. Se ela está bem afastada do zero,
  há um **erro sistemático** entre R1 e R2 (o sistema tende a ler mais alto ou mais
  baixo na repetição).
- As **linhas tracejadas** são os **limites de concordância** (viés ± 1,96·DP);
  espera-se que ~95% dos pontos fiquem dentro delas. Os pontos **fora** aparecem
  listados na tabela logo acima, com o código de barras — são as amostras mais
  discrepantes, candidatas a suspeitas.
- Se a nuvem de pontos **abre como um funil** ou **inclina** conforme a concentração
  aumenta, o erro é **proporcional à concentração** (típico de problema de calibração
  ou linearidade), e não um erro constante.
- Idealmente os pontos ficam espalhados **simetricamente em torno do zero**, sem padrão.

**Média móvel da diferença** — mostra a diferença `R1−R2` de cada amostra na
**ordem cronológica** (quando você informa data e hora) ou na ordem da planilha, com
uma **média móvel** (linha destacada) que suaviza o ruído e revela **tendências ao
longo do tempo/corrida**.

- A média móvel deve **oscilar em torno do zero**. Uma **subida ou descida
  sustentada** indica uma **deriva** do sistema (reagente envelhecendo, calibração
  saindo do lugar, degradação do equipamento) — mesmo que cada par individual pareça
  aceitável.
- Um **degrau/salto abrupto** costuma marcar um **evento**: troca de lote de reagente,
  recalibração, manutenção. Cruzar a posição do salto com o log do equipamento (pela
  data/hora) ajuda a achar a causa.
- Use a **janela** para ajustar a sensibilidade: janela pequena reage rápido a
  mudanças (mais ruído); janela grande evidencia tendências longas (mais suave).
"""
    )

# ---- Bloco 5: exportar ---------------------------------------------------- #
st.markdown("### 5 · Exportar resultados")

# Monta o relatório de saída: remove colunas internas, arredonda, reordena e renomeia.
export = base.drop(columns=["_lo", "_hi", "DataHora", "Suspeito_erro", "Mudou_interp"],
                   errors="ignore").copy()
cols_2dec = [c for c in ["DP_par", "CV_par_%", "ETA_%"]
             if c in export.columns]
for _c in cols_2dec:
    export[_c] = pd.to_numeric(export[_c], errors="coerce").round(2)
# Ordem desejada (nomes internos); o restante segue na ordem atual.
_lead = [c for c in ["ID", "Idade", "Sexo", "R1", "R2", "R1 anterior", "Data R1",
                     "Hora R1", "Equip. R1", "Equip. R2", "RefRange",
                     "Usuário validação R1"] if c in export.columns]
_rest = [c for c in export.columns if c not in _lead]
export = export[_lead + _rest]
export = export.rename(columns={
    "ID": "Código de barras", "R1": "1º Resultado", "R2": "Repetição",
    "R1 anterior": "Resultado anterior", "Equip. R1": "Equipamento R1",
    "Equip. R2": "Equipamento R2",
})

d1, d2 = st.columns(2)
with d1:
    st.download_button("⬇️ Baixar tabela (Excel)", data=to_excel(export, cols_2dec=cols_2dec),
                       file_name="analise_repeticoes.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
with d2:
    csv_bytes = export.to_csv(index=False, sep=";", decimal=",", encoding="utf-8-sig").encode("utf-8-sig")
    st.download_button("⬇️ Baixar tabela (CSV)", data=csv_bytes,
                       file_name="analise_repeticoes.csv", mime="text/csv")
st.caption("O **.xlsx** sai com colunas centralizadas e largura ajustada (autofit). O "
           "**.csv** é texto puro — não guarda largura/alinhamento (isso é do Excel); "
           "ele leva a mesma ordem, nomes e arredondamento das colunas.")
